import os
from openpyxl import Workbook, load_workbook
from basic.generator import generate_people, generate_authors, generate_books
from basic.handler.csv_dict import csv_dict_read, csv_dict_write
from basic.handler.csv_list import read as csv_read, write as csv_write
from basic.handler.my_json import read as json_read, write as json_write
from basic.handler.oracle_db import connect_to_oracle, create_tables, insert_persons, insert_authors, \
    insert_books, fetch_persons, fetch_authors, fetch_books
from basic.handler.xlsx import read as xlsx_read, write as xlsx_write
from basic.model_dataclasses import Person, Author, Book


def main():
    # Elérési út az adatok mentéséhez
    output_path = "output"
    os.makedirs(output_path, exist_ok=True)

    # Adatok generálása
    num_people = 20
    num_authors = 21
    num_books = 22

    # Személyek generálása
    people = generate_people(num_people)
    print("Generated People:", people)

    # Szerzők generálása
    authors = generate_authors(num_authors)
    print("Generated Authors:", authors)

    # Könyvek generálása
    books = generate_books(num_books, authors, people)
    print("Generated Books:", books)

    # Adatok írása CSV fájlokba
    csv_write(people, output_path, "people.csv")
    csv_write(authors, output_path, "authors.csv")
    csv_write(books, output_path, "books.csv")
    print("Data successfully written to CSV files.")

    # Adatok olvasása a CSV fájlokból
    loaded_people_csv = csv_read(Person, output_path, "people.csv")
    loaded_authors_csv = csv_read(Author, output_path, "authors.csv")
    loaded_books_csv = csv_read(Book, output_path, "books.csv")

    # Ellenőrzés: Olvasott adatok kiírása
    print("\nLoaded People from CSV:", loaded_people_csv)
    print("\nLoaded Authors from CSV:", loaded_authors_csv)
    print("\nLoaded Books from CSV:", loaded_books_csv)

    # Adatok írása CSV_dict-be
    csv_dict_write(people, output_path, "people_dict.csv")
    csv_dict_write(authors, output_path, "authors_dict.csv")
    csv_dict_write(books, output_path, "books_dict.csv")
    print("Data successfully written to CSV files using csv_dict_write.")

    # Olvasás CSV_dict-ből
    loaded_people_dict = csv_dict_read(Person, output_path, "people_dict.csv")
    loaded_authors_dict = csv_dict_read(Author, output_path, "authors_dict.csv")
    loaded_books_dict = csv_dict_read(Book, output_path, "books_dict.csv")

    # Ellenőrzés, kiírás
    print("\nLoaded People from Dict CSV:", loaded_people_dict)
    print("\nLoaded Authors from Dict CSV:", loaded_authors_dict)
    print("\nLoaded Books from Dict CSV:", loaded_books_dict)

    # Adatok írása JSON fájlokba
    json_write(people, output_path, "people", ".json")
    json_write(authors, output_path, "authors", ".json")
    json_write(books, output_path, "books", ".json")
    print("Data successfully written to JSON files.")

    # Adatok olvasása a JSON fájlokból
    loaded_people_json = json_read(Person, output_path, "people", ".json")
    loaded_authors_json = json_read(Author, output_path, "authors", ".json")
    loaded_books_json = json_read(Book, output_path, "books", ".json")

    # Ellenőrzés: Olvasott adatok kiírása a konzolra (JSON)
    print("\nLoaded People from JSON:", loaded_people_json)
    print("\nLoaded Authors from JSON:", loaded_authors_json)
    print("\nLoaded Books from JSON:", loaded_books_json)

    # Adatok írása XLSX fájlba
    xlsx_file_path = os.path.join(output_path, "data.xlsx")
    wb = Workbook()
    xlsx_write(people, wb, "people")
    xlsx_write(authors, wb, "authors")
    xlsx_write(books, wb, "books")
    wb.save(xlsx_file_path)
    print("Data successfully written to XLSX file.")

    # Adatok olvasása az XLSX fájlból
    wb = load_workbook(xlsx_file_path)
    loaded_people_xlsx = xlsx_read(Person, wb, "people")
    loaded_authors_xlsx = xlsx_read(Author, wb, "authors")
    loaded_books_xlsx = xlsx_read(Book, wb, "books")

    # Ellenőrzés: Olvasott adatok kiírása a konzolra (XLSX)
    print("\nLoaded People from XLSX:", loaded_people_xlsx)
    print("\nLoaded Authors from XLSX:", loaded_authors_xlsx)
    print("\nLoaded Books from XLSX:", loaded_books_xlsx)

    # Oracle adatbázis kapcsolat létrehozása
    user = "U_C08KB9"
    password = "szantolora"
    host = "codd.inf.unideb.hu"
    port = 1521
    service_name = "ora21cp.inf.unideb.hu"

    connection = connect_to_oracle(user, password, host, port, service_name)

    # Táblák létrehozása az adatbázisban
    create_tables(connection)

    # Adatok beszúrása az adatbázisba
    insert_persons(connection, people)
    insert_authors(connection, authors)
    insert_books(connection, books)
    print("Data successfully inserted into Oracle Database.")

    # Adatok lekérdezése az adatbázisból
    loaded_people_db = fetch_persons(connection)
    loaded_authors_db = fetch_authors(connection)
    loaded_books_db = fetch_books(connection)

    # Ellenőrzés: Olvasott adatok kiírása a konzolra (Oracle Database)
    print("\nLoaded People from Database:", loaded_people_db)
    print("\nLoaded Authors from Database:", loaded_authors_db)
    print("\nLoaded Books from Database:", loaded_books_db)

    # Kapcsolat lezárása
    connection.close()
    print("Oracle Database connection closed.")


if __name__ == "__main__":
    main()
