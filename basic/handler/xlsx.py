from typing import Type, cast
from openpyxl.workbook import Workbook
from basic.model_dataclasses import Person, Author, Book


def write_people(people: list[Person], wb: Workbook, sheet_name: str = "people", heading: bool = True) -> None:
    sheet = wb.create_sheet(sheet_name)

    if heading:
        column_names = ["id", "name", "age", "male"]
        for col in range(len(column_names)):
            sheet.cell(row=1, column=col + 1, value=column_names[col])

    offset = 2 if heading else 1
    for row in range(len(people)):
        sheet.cell(row=row + offset, column=1, value=people[row].id)
        sheet.cell(row=row + offset, column=2, value=people[row].name)
        sheet.cell(row=row + offset, column=3, value=people[row].age)
        sheet.cell(row=row + offset, column=4, value=people[row].male)


def read_people(wb: Workbook, sheet_name: str = "people", heading: bool = True) -> list[Person]:
    sheet = wb[sheet_name]

    people = []
    row = 2 if heading else 1
    while True:
        cell = sheet.cell(row=row, column=1)
        if cell.value is None:
            break

        people.append(
            Person(
                sheet.cell(row=row, column=1).value,
                sheet.cell(row=row, column=2).value,
                sheet.cell(row=row, column=3).value,
                sheet.cell(row=row, column=4).value
            )
        )
        row += 1
    return people


def write_authors(authors: list[Author], wb: Workbook, sheet_name: str = "authors", heading: bool = True) -> None:
    sheet = wb.create_sheet(sheet_name)

    if heading:
        column_names = ["id", "name", "birth_year", "nationality"]
        for col in range(len(column_names)):
            sheet.cell(row=1, column=col + 1, value=column_names[col])

    offset = 2 if heading else 1
    for row in range(len(authors)):
        sheet.cell(row=row + offset, column=1, value=authors[row].id)
        sheet.cell(row=row + offset, column=2, value=authors[row].name)
        sheet.cell(row=row + offset, column=3, value=authors[row].birth_year)
        sheet.cell(row=row + offset, column=4, value=authors[row].nationality)


def read_authors(wb: Workbook, sheet_name: str = "authors", heading: bool = True) -> list[Author]:
    sheet = wb[sheet_name]

    authors = []
    row = 2 if heading else 1
    while True:
        cell = sheet.cell(row=row, column=1)
        if cell.value is None:
            break

        authors.append(
            Author(
                sheet.cell(row=row, column=1).value,
                sheet.cell(row=row, column=2).value,
                sheet.cell(row=row, column=3).value,
                sheet.cell(row=row, column=4).value
            )
        )
        row += 1
    return authors


def write_books(books: list[Book], wb: Workbook, sheet_name: str = "books", heading: bool = True) -> None:
    sheet = wb.create_sheet(sheet_name)

    if heading:
        column_names = ["id", "title", "genre", "publication_year", "author_id", "person_id"]
        for col in range(len(column_names)):
            sheet.cell(row=1, column=col + 1, value=column_names[col])

    offset = 2 if heading else 1
    for row in range(len(books)):
        sheet.cell(row=row + offset, column=1, value=books[row].id)
        sheet.cell(row=row + offset, column=2, value=books[row].title)
        sheet.cell(row=row + offset, column=3, value=books[row].genre)
        sheet.cell(row=row + offset, column=4, value=books[row].publication_year)
        sheet.cell(row=row + offset, column=5, value=books[row].author_id)
        sheet.cell(row=row + offset, column=6, value=books[row].person_id)


def read_books(wb: Workbook, sheet_name: str = "books", heading: bool = True) -> list[Book]:
    sheet = wb[sheet_name]

    books = []
    row = 2 if heading else 1
    while True:
        cell = sheet.cell(row=row, column=1)
        if cell.value is None:
            break

        books.append(
            Book(
                sheet.cell(row=row, column=1).value,
                sheet.cell(row=row, column=2).value,
                sheet.cell(row=row, column=3).value,
                sheet.cell(row=row, column=4).value,
                sheet.cell(row=row, column=5).value,
                sheet.cell(row=row, column=6).value
            )
        )
        row += 1
    return books


def write(entities: list[object], wb: Workbook, sheet_name: str = None, heading: bool = True) -> None:
    if isinstance(entities[0], Person):
        return write_people([cast(Person, e) for e in entities], wb, sheet_name=sheet_name, heading=heading)
    elif isinstance(entities[0], Author):
        return write_authors([cast(Author, e) for e in entities], wb, sheet_name=sheet_name, heading=heading)
    elif isinstance(entities[0], Book):
        return write_books([cast(Book, e) for e in entities], wb, sheet_name=sheet_name, heading=heading)
    else:
        raise RuntimeError("Unknown type of entity")


def read(entity_type: Type[object], wb: Workbook, sheet_name: str = None, heading: bool = True) -> list[object]:
    if entity_type == Person:
        return read_people(wb, sheet_name=sheet_name, heading=heading)
    elif entity_type == Author:
        return read_authors(wb, sheet_name=sheet_name, heading=heading)
    elif entity_type == Book:
        return read_books(wb, sheet_name=sheet_name, heading=heading)
    else:
        raise RuntimeError("Unknown type of entity")
